#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generuje assets/js/machines-data.js i machines-data-en.js z arkusza
referencja_nagran_maszyny.xlsx.

Odpowiednik generate_space_data.py dla serii Maszyny Budowlane. Roznice:

1. Atrybucja jest OBOWIAZKOWA przy 14 z 40 nagran (licencja CC BY), wiec
   pole attribution nigdy nie jest puste i zawsze ma link do licencji.
2. Karta pokazuje dodatkowo opis dzwieku, bo nagranie pochodzi z maszyny
   tego samego rodzaju, nie zawsze z egzemplarza narysowanego.
3. Slug jest wspolny dla obu jezykow, tak jak przy Kosmosie: kody QR w obu
   wydaniach maja ten sam hash, rozni je tylko adres strony.

UZYCIE:
    python3 generate_machines_data.py
"""
import json, os, sys
try:
    import openpyxl
except ImportError:
    sys.exit("Brak openpyxl: pip install openpyxl --break-system-packages")

TU = os.path.dirname(os.path.abspath(__file__))
ARKUSZ = os.path.join(TU, "referencja_nagran_maszyny.xlsx")
AUDIO = os.path.join(TU, "audio", "maszyny")
GRAFIKI = os.path.join(TU, "images", "maszyny")
WY_PL = os.path.join(TU, "assets", "js", "machines-data.js")
WY_EN = os.path.join(TU, "assets", "js", "machines-data-en.js")

NAGLOWEK = """/* Rejestr stron serii Maszyny Budowlane - WYGENEROWANY AUTOMATYCZNIE
   przez generate_machines_data.py. Nie edytuj recznie.
   Zeby cos zmienic: popraw referencja_nagran_maszyny.xlsx i uruchom skrypt. */

const MACHINES = [
"""


def wczytaj():
    ws = openpyxl.load_workbook(ARKUSZ, data_only=True).active
    nag = [c.value for c in ws[1]]
    idx = {n: i for i, n in enumerate(nag)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["Slug"]]:
            continue
        out.append({k: row[i] for k, i in idx.items()})
    return out


def blok(w, jezyk, prefiks):
    nazwa = w["Nazwa EN"] if jezyk == "en" else w["Nazwa PL"]
    fakt = w["Ciekawostka EN"] if jezyk == "en" else w["Ciekawostka PL"]
    dzwiek = w["Opis dzwieku EN"] if jezyk == "en" else w["Opis dzwieku PL"]
    autor = w["Autor nagrania"]
    lic = w["Licencja"]
    atryb = (f"Recording: {autor}, freesound.org ({lic})" if jezyk == "en"
             else f"Nagranie: {autor}, freesound.org ({lic})")
    return (
        "  {\n"
        f'    slug: {json.dumps(w["Slug"])},\n'
        f'    name: {json.dumps(nazwa)},\n'
        f'    fact: {json.dumps(fakt)},\n'
        f'    sound: {json.dumps(dzwiek)},\n'
        f'    audioUrl: {json.dumps(prefiks + "audio/maszyny/" + w["Slug"] + ".mp3")},\n'
        f'    imageUrl: {json.dumps(prefiks + "images/maszyny/" + w["Slug"] + ".webp")},\n'
        f'    attribution: {json.dumps(atryb)},\n'
        f'    licenseUrl: {json.dumps(w["Link do licencji"])},\n'
        f'    sourceUrl: {json.dumps(w["Strona zrodla"])},\n'
        "  },\n"
    )


def main():
    wiersze = wczytaj()
    dobre, braki = [], []
    for w in wiersze:
        s = w["Slug"]
        a = os.path.exists(os.path.join(AUDIO, s + ".mp3"))
        g = os.path.exists(os.path.join(GRAFIKI, s + ".webp"))
        (dobre if (a and g) else braki).append((w, a, g))
    for wy, jez in ((WY_PL, "pl"), (WY_EN, "en")):
        with open(wy, "w", encoding="utf-8") as f:
            f.write(NAGLOWEK)
            for w, _, _ in dobre:
                f.write(blok(w, jez, "../"))
            f.write("];\n")
        print(f"zapisano {wy}: {len(dobre)} kart")
    if braki:
        print("\nPOMINIETE:")
        for w, a, g in braki:
            print(f"  {w['Slug']}: brakuje", "audio" if not a else "", "grafiki" if not g else "")
    else:
        print("Wszystkie 40 stron maja audio i grafike.")


if __name__ == "__main__":
    main()
