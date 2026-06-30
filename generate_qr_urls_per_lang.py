"""
Generuje 4 pliki qr_urls_{lang}.xlsx (PL, EN, DE, ES) w identycznym formacie
co istniejacy qr_urls.xlsx (kolumny: "Gatunek (PL)", "Slug", "URL do QR kodu"),
do wklejenia w czacie generujacym karty ptakow z kodami QR.

Kolumna "Gatunek (PL)" jest STALYM kluczem mapowania - identyczna we
wszystkich 4 plikach, zawsze polska nazwa gatunku (zgodnie z konwencja
tamtego projektu: "nazwa_pl_ref jest wymagana jako klucz do mapowania na
plik obrazu", patrz PODSUMOWANIE_SESJI.md). Kolumny "Slug" i "URL do QR
kodu" sa PRZETLUMACZONE per jezyk: Slug = localSlug (np. niemieckie
"fasan", nie polskie "bazant"), URL wskazuje na wlasciwa podstrone
jezykowa media.bartechbooks.com.

Source of truth dla par (slug PL, localSlug, nazwa wyswietlana):
assets/js/birds-data.js (PL, slug==localSlug) i birds-data-{en,de,es}.js.

Sortowanie: alfabetycznie po localSlug, dla EN/DE/ES. To jest rownowazne
sortowaniu language-aware po nazwie wyswietlanej (np. niemieckie
Bluthänfling -> localSlug "bluthaenfling" siada poprawnie w sekcji B, nie
po Z), bo localSlug JEST JUZ transliterowana, zlatynizowana wersja nazwy.
DLA PL: plik nie jest generowany na nowo - kopiowany 1:1 z istniejacego,
juz sprawdzonego qr_urls.xlsx. Powod: sortowanie po localSlug (ASCII po
usunieciu polskich znakow: ą,ć,ę,ł,ń,ó,ś,ź,ż -> najblizsza litera bez
ogonka) NIE jest rownowazne prawdziwemu polskiemu porzadkowi alfabetycznemu
- w prawdziwym polskim 'ą' jest osobna litera PO 'a', wiec np. "Gawron"
(Ga-) powinno wyjsc PRZED "Gągoł" (Gą-), ale localSlug zamienia obie na
"ga-" i ASCII sortuje je inaczej (po drugiej literze: gagol < gawron).
To jest realna, wykryta w tej sesji rozbieznosc, nie teoretyczna - PL ma
juz wlasny, zweryfikowany porzadek w qr_urls.xlsx, nie psuc go.

Uzycie: python3 generate_qr_urls_per_lang.py
Wyjscie: qr_urls_pl.xlsx (kopia oryginalu), qr_urls_en.xlsx, qr_urls_de.xlsx,
qr_urls_es.xlsx (nowo wygenerowane, sortowane po localSlug).
"""
import re
import json
import shutil
import openpyxl

BASE_URL = "https://media.bartechbooks.com"

PL_SOURCE_FILE = "assets/js/birds-data.js"
EXISTING_PL_XLSX = "qr_urls.xlsx"  # juz istnieje, sprawdzony porzadek - kopiowany 1:1

LANGS = {
    "en": {"file": "assets/js/birds-data-en.js", "path": "en/birds"},
    "de": {"file": "assets/js/birds-data-de.js", "path": "de/voegel"},
    "es": {"file": "assets/js/birds-data-es.js", "path": "es/aves"},
}


def parse_birds_js(path):
    """Parsuje plik birds-data*.js, zwraca liste dictow {slug, localSlug, namePl}."""
    content = open(path, encoding="utf-8").read()
    pattern = re.compile(
        r'slug:\s*"([^"]+)",\s*'
        r'localSlug:\s*"([^"]+)",\s*'
        r'namePl:\s*"((?:[^"\\]|\\.)*)"'
    )
    out = []
    for slug, local_slug, name_raw in pattern.findall(content):
        name = json.loads('"' + name_raw + '"')
        out.append({"slug": slug, "localSlug": local_slug, "namePl": name})
    return out


def main():
    # PL: kopia istniejacego, juz sprawdzonego pliku - nie regenerujemy,
    # zobacz wyjasnienie w docstringu modulu (problem z sortowaniem ASCII
    # po usunieciu polskich znakow diakrytycznych).
    shutil.copy(EXISTING_PL_XLSX, "qr_urls_pl.xlsx")
    wb_check = openpyxl.load_workbook("qr_urls_pl.xlsx")
    n_pl = wb_check.active.max_row - 1
    print(f"qr_urls_pl.xlsx: {n_pl} wierszy (kopia {EXISTING_PL_XLSX}, bez zmian)")

    # Polska nazwa per slug - klucz mapowania dla EN/DE/ES, brany WYLACZNIE
    # z rejestru PL.
    pl_birds = parse_birds_js(PL_SOURCE_FILE)
    pl_name_by_slug = {b["slug"]: b["namePl"] for b in pl_birds}

    for lang, cfg in LANGS.items():
        birds = parse_birds_js(cfg["file"])
        birds.sort(key=lambda b: b["localSlug"])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "URL do QR kodow"
        ws.append(["Gatunek (PL)", "Slug", "URL do QR kodu"])

        for b in birds:
            gatunek_pl = pl_name_by_slug[b["slug"]]  # stale, zawsze polskie
            local_slug = b["localSlug"]
            url = f"{BASE_URL}/{cfg['path']}/#{local_slug}"
            ws.append([gatunek_pl, local_slug, url])

        out_name = f"qr_urls_{lang}.xlsx"
        wb.save(out_name)
        print(f"{out_name}: {len(birds)} wierszy")


if __name__ == "__main__":
    main()
