#!/usr/bin/env python3
"""
Generuje assets/js/birds-data.js na podstawie:
  - plikow audio w assets/audio/ptaki/*.mp3
  - plikow grafik w assets/images/ptaki/*.png (lub .jpg)
  - metadanych (autor, licencja, nazwa lacinska) z referencja_nagran.xlsx

Logika dopasowania: slug pliku (nazwa bez rozszerzenia) musi byc identyczny
miedzy audio i grafika, oraz pokrywac sie z kolumna 'Plik' w arkuszu
(po usunieciu rozszerzenia .mp3) zeby dociagnac atrybucje.

Gatunki bez pliku audio LUB bez grafiki sa pomijane w wygenerowanym rejestrze
i wypisywane w raporcie - karta bez dzwieku albo bez obrazka nie powinna trafic
na strone niedokonczona.

UZYCIE:
    python3 generate_birds_data.py
    python3 generate_birds_data.py --audio-dir inna/sciezka --images-dir inna/sciezka
"""

import argparse
import json
import os
import subprocess
import sys
import unicodedata

try:
    import openpyxl
except ImportError:
    print("Brak openpyxl. Zainstaluj: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)


PL_MAP = str.maketrans({
    "ą": "a", "ć": "c", "ę": "e", "ł": "l", "ń": "n", "ó": "o", "ś": "s", "ź": "z", "ż": "z",
    "Ą": "A", "Ć": "C", "Ę": "E", "Ł": "L", "Ń": "N", "Ó": "O", "Ś": "S", "Ź": "Z", "Ż": "Z",
})


def slugify(name: str) -> str:
    """Normalizuje nazwe pliku do sluga: usuwa polskie znaki, lowercase,
    zamienia spacje i podkreslniki na myslniki. Uzywane konsekwentnie dla
    NAZW AUDIO i NAZW GRAFIK, zeby dopasowanie dzialalo niezaleznie od
    konwencji nazewnictwa kazdej dostawy plikow (wielkie/male litery,
    podkreslniki vs myslniki)."""
    name = name.translate(PL_MAP)
    nfkd = unicodedata.normalize("NFKD", name)
    ascii_name = "".join(c for c in nfkd if not unicodedata.combining(c))
    return ascii_name.lower().replace(" ", "-").replace("_", "-")


def load_reference(xlsx_path: str) -> dict:
    """Wczytuje arkusz referencyjny, zwraca dict slug -> metadane."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fname = row[idx["Plik"]]
        if not fname:
            continue
        slug = fname[:-4] if fname.endswith(".mp3") else fname
        out[slug] = {
            "nameLat": row[idx["Gatunek (lat.)"]],
            "namePl": row[idx["Gatunek (PL)"]],
            "author": row[idx["Autor / nagrywajacy"]],
            "license": row[idx["Licencja"]],
            "sourceUrl": row[idx["Link zrodlowy"]],
            "sourceName": row[idx["Zrodlo"]],
        }
    return out


LICENSE_URLS = {
    "cc0": "https://creativecommons.org/publicdomain/zero/1.0/",
    "by": "https://creativecommons.org/licenses/by/4.0/",
    "by-sa": "https://creativecommons.org/licenses/by-sa/4.0/",
}


SOURCE_DISPLAY_NAMES = {
    "xeno-canto": "xeno-canto.org",
    "Wikimedia Commons": "Wikimedia Commons",
}


def check_audio_codec(path: str) -> str | None:
    """Sprawdza realny kodek pliku audio przez ffprobe. Zwraca nazwe kodeka
    albo None jesli ffprobe nie jest dostepne/plik nieprawidlowy.
    Uzywane do wykrycia plikow .mp3 ktore w rzeczywistosci sa surowym PCM/WAV
    (zdarzylo sie to dla ~20 plikow z xeno-canto, ktorych URL mial rozszerzenie
    .mp3 mimo ze realny format byl inny - przegladarka nie moze seekowac w
    takich plikach mimo poprawnego odtwarzania)."""
    try:
        result = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "stream=codec_name",
             "-of", "default=noprint_wrappers=1:nokey=1", path],
            capture_output=True, text=True, timeout=10,
        )
        return result.stdout.strip() or None
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--audio-dir", default="audio/ptaki")
    parser.add_argument("--images-dir", default="images/ptaki")
    parser.add_argument("--reference", default="referencja_nagran.xlsx")
    parser.add_argument("--out", default="assets/js/birds-data.js")
    args = parser.parse_args()

    audio_files = {}
    for f in os.listdir(args.audio_dir):
        if f.endswith(".mp3"):
            slug = slugify(f[:-4])
            if slug != f[:-4]:
                print(f"  [normalizacja] audio '{f}' -> slug '{slug}'")
            audio_files[slug] = f
    audio_slugs = set(audio_files.keys())

    image_files = {}
    for f in os.listdir(args.images_dir):
        for ext in (".png", ".jpg", ".jpeg", ".webp"):
            if f.lower().endswith(ext):
                base = f[:-len(ext)]
                slug = slugify(base)
                image_files[slug] = f
                break

    reference = {}
    if os.path.exists(args.reference):
        reference = load_reference(args.reference)
    else:
        print(f"UWAGA: nie znaleziono '{args.reference}', rejestr bedzie bez atrybucji.", file=sys.stderr)

    both = audio_slugs & set(image_files.keys())
    only_audio = audio_slugs - set(image_files.keys())
    only_image = set(image_files.keys()) - audio_slugs

    if only_audio:
        print(f"POMIJAM (brak grafiki): {sorted(only_audio)}")
    if only_image:
        print(f"POMIJAM (brak audio): {sorted(only_image)}")

    no_reference = [s for s in both if s not in reference]
    if no_reference:
        print(f"UWAGA - brak atrybucji w arkuszu dla: {sorted(no_reference)} (karta bedzie bez autora/licencji)")

    birds = []
    bad_codec_files = []
    for slug in sorted(both):
        ref = reference.get(slug, {})
        license_code = ref.get("license")

        audio_path = os.path.join(args.audio_dir, audio_files[slug])
        codec = check_audio_codec(audio_path)
        if codec and codec != "mp3":
            bad_codec_files.append((slug, codec))

        birds.append({
            "slug": slug,
            "localSlug": slug,  # PL: slug juz jest poprawna forma URL-safe polskiej nazwy,
                                  # localSlug istnieje rownolegle bo rejestry EN/DE/ES uzywaja
                                  # przetlumaczonej, jezykowo-specyficznej wersji (zobacz
                                  # assets/js/inject_local_slug.py) - card.id w player.js
                                  # zawsze czyta bird.localSlug, niezaleznie od jezyka.
            "namePl": ref.get("namePl") or slug.replace("-", " ").title(),
            "nameLat": ref.get("nameLat") or "",
            "audioUrl": f"../audio/ptaki/{audio_files[slug]}",
            "imageUrl": f"../images/ptaki/{image_files[slug]}",
            "attribution": f"Nagranie: {ref['author']}, {SOURCE_DISPLAY_NAMES.get(ref.get('sourceName'), ref.get('sourceName') or 'xeno-canto.org')}" if ref.get("author") else None,
            "licenseUrl": LICENSE_URLS.get(license_code, ref.get("sourceUrl")),
        })

    if bad_codec_files:
        print(f"\nUWAGA KRYTYCZNA - {len(bad_codec_files)} plikow .mp3 ma w rzeczywistosci inny kodek")
        print("(seek/przewijanie NIE bedzie dzialac w przegladarce dla tych plikow):")
        for slug, codec in bad_codec_files:
            print(f"  {slug}: realny kodek = {codec}")
        print("Naprawa: ffmpeg -i plik.mp3 -codec:a libmp3lame -b:a 192k plik_fixed.mp3")

    js_lines = ["/* Rejestr ptakow - WYGENEROWANY AUTOMATYCZNIE przez generate_birds_data.py.",
                "   Nie edytuj recznie - zmiany przepadna przy nastepnym uruchomieniu skryptu.",
                "   Zeby dodac/zmienic ptaka: dodaj pliki audio+grafika, uzupelnij arkusz",
                "   referencyjny, uruchom skrypt ponownie. */", "", "const BIRDS = ["]
    for b in birds:
        js_lines.append("  {")
        js_lines.append(f"    slug: {json.dumps(b['slug'], ensure_ascii=False)},")
        js_lines.append(f"    localSlug: {json.dumps(b['localSlug'], ensure_ascii=False)},")
        js_lines.append(f"    namePl: {json.dumps(b['namePl'], ensure_ascii=False)},")
        js_lines.append(f"    nameLat: {json.dumps(b['nameLat'], ensure_ascii=False)},")
        js_lines.append(f"    audioUrl: {json.dumps(b['audioUrl'])},")
        js_lines.append(f"    imageUrl: {json.dumps(b['imageUrl'])},")
        js_lines.append(f"    attribution: {json.dumps(b['attribution'], ensure_ascii=False) if b['attribution'] else 'null'},")
        js_lines.append(f"    licenseUrl: {json.dumps(b['licenseUrl']) if b['licenseUrl'] else 'null'},")
        js_lines.append("  },")
    js_lines.append("];")

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write("\n".join(js_lines) + "\n")

    print(f"\nWygenerowano {len(birds)} ptakow -> {args.out}")


if __name__ == "__main__":
    main()
