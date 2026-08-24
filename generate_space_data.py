#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generuje assets/js/space-data.js (PL) i assets/js/space-data-en.js (EN)
na podstawie referencja_nagran_kosmos.xlsx.

Odpowiednik generate_birds_data.py dla serii Kosmos. Roznice wzgledem ptakow:

1. Slug jest WSPOLNY dla obu jezykow. Kody QR w obu wydaniach ksiazki
   prowadza do tego samego hasha (np. #start-rakiety), tylko na inna
   strone: /kosmos/ albo /space/. Nie ma odpowiednika localSlug.
2. Karta pokazuje ciekawostke (pelne zdanie), a nie nazwe lacinska.
3. Zrodlo jest opisane misja plus instytucja, bez linku do pliku.

Sprawdza, ze kazdy slug ma plik audio i grafike. Brakujace pomija
i wypisuje w raporcie - karta bez dzwieku nie powinna trafic na strone.

UZYCIE:
    python3 generate_space_data.py
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Brak openpyxl. Zainstaluj: pip install openpyxl --break-system-packages",
          file=sys.stderr)
    sys.exit(1)

TU = os.path.dirname(os.path.abspath(__file__))
ARKUSZ = os.path.join(TU, "referencja_nagran_kosmos.xlsx")
AUDIO = os.path.join(TU, "audio", "kosmos")
GRAFIKI = os.path.join(TU, "images", "kosmos")
WY_PL = os.path.join(TU, "assets", "js", "space-data.js")
WY_EN = os.path.join(TU, "assets", "js", "space-data-en.js")

NAGLOWEK = """/* Rejestr stron serii Kosmos - WYGENEROWANY AUTOMATYCZNIE przez
   generate_space_data.py. Nie edytuj recznie - zmiany przepadna przy
   nastepnym uruchomieniu skryptu.
   Zeby cos zmienic: popraw referencja_nagran_kosmos.xlsx i uruchom skrypt. */

const SPACE = [
"""


def wczytaj_arkusz(sciezka):
    wb = openpyxl.load_workbook(sciezka)
    ws = wb.active
    naglowek = [c.value for c in ws[1]]
    idx = {n: i for i, n in enumerate(naglowek)}
    wiersze = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["Slug"]]:
            continue
        wiersze.append({k: row[i] for k, i in idx.items()})
    return wiersze


def blok(w, jezyk, prefiks):
    nazwa = w["Nazwa EN"] if jezyk == "en" else w["Nazwa PL"]
    fakt = w["Ciekawostka EN"] if jezyk == "en" else w["Ciekawostka PL"]
    nagr = w["Nagranie EN"] if jezyk == "en" else w["Nagranie PL"]
    zrodlo = ("Recording: " if jezyk == "en" else "Nagranie: ") + \
             f'{w["Misja"]}, {w["Instytucja"]}'
    return (
        "  {\n"
        f'    slug: {json.dumps(w["Slug"])},\n'
        f'    name: {json.dumps(nazwa)},\n'
        f'    fact: {json.dumps(fakt)},\n'
        f'    sound: {json.dumps(nagr)},\n'
        f'    audioUrl: {json.dumps(prefiks + "audio/kosmos/" + w["Slug"] + ".mp3")},\n'
        f'    imageUrl: {json.dumps(prefiks + "images/kosmos/" + w["Slug"] + ".webp")},\n'
        f'    attribution: {json.dumps(zrodlo)},\n'
        "  },\n"
    )


def main():
    wiersze = wczytaj_arkusz(ARKUSZ)
    braki = []
    dobre = []
    for w in wiersze:
        s = w["Slug"]
        ma_audio = os.path.exists(os.path.join(AUDIO, s + ".mp3"))
        ma_grafike = os.path.exists(os.path.join(GRAFIKI, s + ".webp"))
        if ma_audio and ma_grafike:
            dobre.append(w)
        else:
            braki.append((s, "audio" if not ma_audio else "", "grafika" if not ma_grafike else ""))

    # Obie strony (kosmos/ i space/) leza jeden poziom pod korzeniem,
    # wiec prefiks jest ten sam. Ptaki mialy en/birds/, stad tam ../../.
    for wy, jezyk in ((WY_PL, "pl"), (WY_EN, "en")):
        with open(wy, "w", encoding="utf-8") as f:
            f.write(NAGLOWEK)
            for w in dobre:
                f.write(blok(w, jezyk, "../"))
            f.write("];\n")
        print(f"zapisano {wy}: {len(dobre)} kart")

    if braki:
        print("\nPOMINIETE (brak pliku):")
        for s, a, g in braki:
            print(f"  {s}: brakuje {' i '.join(x for x in (a, g) if x)}")
    else:
        print("Wszystkie 40 stron maja audio i grafike.")


if __name__ == "__main__":
    main()
